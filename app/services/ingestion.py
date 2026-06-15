"""
Document text extraction service.
Supports: PDF (PyMuPDF), DOCX (python-docx), XLSX (openpyxl), TXT, MD.
"""
import io
from pathlib import Path


def extract_text(file_bytes: bytes, filename: str) -> str:
    """Return plain text extracted from the uploaded file."""
    ext = Path(filename).suffix.lower().lstrip(".")

    if ext == "pdf":
        return _extract_pdf(file_bytes)
    elif ext == "docx":
        return _extract_docx(file_bytes)
    elif ext in ("xlsx", "xls"):
        return _extract_xlsx(file_bytes)
    elif ext in ("txt", "md", "csv"):
        return file_bytes.decode("utf-8", errors="replace")
    else:
        # Best-effort UTF-8 for unknown types
        return file_bytes.decode("utf-8", errors="replace")


def _extract_pdf(data: bytes) -> str:
    import fitz  # PyMuPDF

    doc = fitz.open(stream=data, filetype="pdf")
    pages = []
    for page in doc:
        pages.append(page.get_text())
    return "\n\n".join(pages)


def _extract_docx(data: bytes) -> str:
    from docx import Document

    doc = Document(io.BytesIO(data))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs)


def _extract_xlsx(data: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
    rows = []
    for sheet in wb.worksheets:
        rows.append(f"## Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            row_text = "\t".join(str(c) if c is not None else "" for c in row)
            if row_text.strip():
                rows.append(row_text)
    return "\n".join(rows)


SUPPORTED_EXTENSIONS = {"pdf", "docx", "xlsx", "xls", "txt", "md", "csv"}


def infer_file_type(filename: str) -> str:
    ext = Path(filename).suffix.lower().lstrip(".")
    return ext if ext in SUPPORTED_EXTENSIONS else "txt"
